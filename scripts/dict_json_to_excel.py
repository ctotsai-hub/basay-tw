#!/usr/bin/env python3
"""
dict_json_to_excel.py
=====================
Regenerate the master Excel file from per-letter JSON entries.

INPUT
-----
  dictionary/entries/*.json
  dictionary/categories.json   (used to restore full category labels)

OUTPUT
------
  dictionary/source/basay_dictionary.xlsm  (or path given via --output)

  The output has the same column layout as the input Excel:
    ID | basay | pos (category, full label) | zh | ja | en
       | pos (source code) | original_entry | remark

  Multi-value cells use '|' as a separator inside Excel (zh/ja/en).
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT       = Path(__file__).resolve().parent.parent
ENTRIES_DIR     = REPO_ROOT / "dictionary" / "entries"
CATEGORIES_JSON = REPO_ROOT / "dictionary" / "categories.json"
DEFAULT_OUT     = REPO_ROOT / "dictionary" / "source" / "basay_dictionary.xlsm"

# Excel column layout (left-to-right). Logical name -> Excel header text.
COLUMNS = [
    ("id",       "ID"),
    ("basay",    "basay"),
    ("category", "pos"),       # full label, e.g. "29情緒思維（精神性）"
    ("zh",       "zh"),
    ("ja",       "ja"),
    ("en",       "en"),
    ("source",   "pos"),       # source code (B/T/M/S/V) - duplicated header
    ("original", "original_entry"),
    ("remark",   "remark"),
]
MULTIVALUE_COLUMNS = {"zh", "ja", "en"}

WIDTHS = {
    "id":       7,
    "basay":    24,
    "category": 24,
    "zh":       30,
    "ja":       30,
    "en":       30,
    "source":   8,
    "original": 22,
    "remark":   38,
}


def load_entries() -> list[dict[str, Any]]:
    if not ENTRIES_DIR.exists():
        sys.exit(f"ERROR: entries directory not found: {ENTRIES_DIR}")
    all_entries: list[dict[str, Any]] = []
    files = sorted(p for p in ENTRIES_DIR.glob("*.json")
                   if p.name != "_index.json")
    if not files:
        sys.exit(f"ERROR: no letter JSON files under {ENTRIES_DIR}")
    for p in files:
        with p.open(encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list):
            sys.exit(f"ERROR: {p} is not a JSON array")
        all_entries.extend(data)
    all_entries.sort(key=lambda e: e.get("id", ""))
    return all_entries


def load_category_map() -> dict[str, str]:
    if CATEGORIES_JSON.is_file():
        with CATEGORIES_JSON.open(encoding="utf-8") as f:
            return json.load(f)
    return {}


def write_excel(entries: list[dict[str, Any]], categories: dict[str, str],
                out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "basay_dictionary"

    headers = [hdr for _, hdr in COLUMNS]
    ws.append(headers)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEEEEE")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")

    wrap = Alignment(wrap_text=True, vertical="top")

    for entry in entries:
        row: list[str] = []
        for logical, _ in COLUMNS:
            if logical == "category":
                num = entry.get("category", "")
                value = categories.get(num, num)  # restore full label
            else:
                value = entry.get(logical, "")
            if logical in MULTIVALUE_COLUMNS and isinstance(value, list):
                value = " | ".join(str(v) for v in value)
            row.append("" if value is None else str(value))
        ws.append(row)

    # Wrap text in multi-value columns + remark.
    for r in range(2, ws.max_row + 1):
        for i, (logical, _) in enumerate(COLUMNS, start=1):
            if logical in MULTIVALUE_COLUMNS or logical == "remark":
                ws.cell(row=r, column=i).alignment = wrap

    for i, (logical, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(logical, 15)

    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    # Save as .xlsx if extension isn't .xlsm (openpyxl can't write .xlsm with VBA).
    if out_path.suffix.lower() == ".xlsm":
        out_path = out_path.with_suffix(".xlsx")
        print(f"  note: writing .xlsx (openpyxl cannot preserve VBA macros): {out_path}")
    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--output", "-o", type=Path, default=DEFAULT_OUT,
                        help=f"output Excel path (default: {DEFAULT_OUT})")
    args = parser.parse_args()

    entries = load_entries()
    categories = load_category_map()
    print(f"Loaded {len(entries)} entries; {len(categories)} categories")
    write_excel(entries, categories, args.output)
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
