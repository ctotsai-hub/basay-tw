#!/usr/bin/env python3
"""
Quick diagnostic: print raw cell values openpyxl reads from the Excel.
Usage:
  python scripts/_debug_excel.py              -- prints first 20 rows
  python scripts/_debug_excel.py 0417 0025    -- prints specific IDs
"""
import sys
from pathlib import Path
import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX = REPO_ROOT / "dictionary" / "source" / "basay_dictionary.xlsm"

ids = set(x.zfill(4) for x in sys.argv[1:]) if len(sys.argv) > 1 else None

wb = openpyxl.load_workbook(XLSX, data_only=True,
                             keep_vba=True)
ws = wb["basay_dictionary"] if "basay_dictionary" in wb.sheetnames else wb.active

rows = ws.iter_rows(values_only=True)
header = next(rows)
print("Header:", header)
print()

count = 0
for row in rows:
    if all(v is None for v in row):
        continue
    raw_id = str(row[0]).strip() if row[0] is not None else ""
    zid = raw_id.zfill(4) if raw_id.isdigit() else raw_id
    if ids:
        if zid not in ids:
            continue
    else:
        if count >= 20:
            break
    print(f"ID={zid} | basay={row[1]!r} | cat={row[2]!r} | zh={row[3]!r} | ja={row[4]!r} | en={row[5]!r}")
    count += 1

wb.close()
