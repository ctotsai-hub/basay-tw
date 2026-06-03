#!/usr/bin/env python3
"""
dict_excel_to_json.py
=====================
Convert basay_dictionary.xlsm into per-letter JSON files (Plan B) plus
merged outputs used by the site.

Expected Excel columns (basay_dictionary sheet):
  ID, basay, pos, zh, ja, en, pos, original_entry, [remark]

Note: TWO columns are named "pos". They are disambiguated by position:
  - first  pos -> "category" (semantic category, leading digit kept)
  - second pos -> "source"   (B/T/M/S/V variety marker)

Outputs:
  dictionary/entries/<letter>.json    per-letter (sorted by basay)
  dictionary/entries/_index.json      id -> letter mapping
  dictionary/categories.json          "29" -> "29情緒思維（精神性）"  (from pos_summary sheet)
  data/dictionary.json                merged list (site-facing)
  data/search-index.json              lightweight search index

Audio detection (does not generate):
  If dictionary/audio/{ipay,hokkien}/<slug>.mp3 exists, an "audio" field
  is attached to the entry. Generation lives in dict_build_audio.py.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
try:
    import basay_text  # type: ignore
    _HAS_BASAY_TEXT = True
except Exception:
    _HAS_BASAY_TEXT = False


REPO_ROOT       = Path(__file__).resolve().parent.parent
DEFAULT_INPUT   = REPO_ROOT / "dictionary" / "source" / "basay_dictionary.xlsm"
ENTRIES_DIR     = REPO_ROOT / "dictionary" / "entries"
CATEGORIES_JSON = REPO_ROOT / "dictionary" / "categories.json"
SITE_JSON       = REPO_ROOT / "data" / "dictionary.json"
SEARCH_INDEX    = REPO_ROOT / "data" / "search-index.json"

AUDIO_ROOT     = REPO_ROOT / "dictionary" / "audio"
AUDIO_VARIANTS = ("ipay", "hokkien")
AUDIO_EXT      = ".mp3"

# Multi-value separators for zh/ja/en (Excel convention):
#   '|'      newline (Alt+Enter)   Chinese comma (、)   Chinese semicolon (；)
MULTIVALUE_SPLIT_RE = re.compile(r"\s*[|\n\r、；]+\s*")

# Slug normalization fallback when basay_text isn't available.
_SIMPLE_SLUG_RE = re.compile(r"[^a-z0-9]+")

# Category prefix: "29情緒思維（精神性）" -> "29"
_CATEGORY_PREFIX_RE = re.compile(r"^\s*(\d{1,3})")

ID_WIDTH = 4   # zero-padded ID width


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    return str(value).strip()


def split_multivalue(cell: str) -> list[str]:
    if not cell:
        return []
    parts = MULTIVALUE_SPLIT_RE.split(cell)
    return [p.strip() for p in parts if p.strip()]


def initial_letter(basay: str) -> str:
    for ch in basay:
        low = ch.lower()
        if "a" <= low <= "z":
            return low
    return "_misc"


def derive_slug(basay: str) -> str:
    if not basay:
        return ""
    # If basay contains "|" variants, use the first one.
    primary = basay.split("|")[0].strip()
    if _HAS_BASAY_TEXT:
        try:
            return basay_text.derive(primary)["slug"]
        except Exception:
            pass
    return _SIMPLE_SLUG_RE.sub("_", primary.lower()).strip("_")


def _slug_letter(slug: str) -> str:
    """First ASCII letter of slug, or '_misc'. Used for audio subfolder bucketing."""
    if slug:
        first = slug[0].lower()
        if "a" <= first <= "z":
            return first
    return "_misc"


def detect_audio(slug: str) -> dict[str, str] | None:
    """Look for MP3 files in letter subfolders, with fallback to flat layout
    for backward compatibility with the pre-migration audio directory."""
    if not slug:
        return None
    letter = _slug_letter(slug)
    found: dict[str, str] = {"slug": slug}
    for variant in AUDIO_VARIANTS:
        # Preferred: dictionary/audio/<variant>/<letter>/<slug>.mp3
        rel = Path("dictionary") / "audio" / variant / letter / f"{slug}{AUDIO_EXT}"
        if (REPO_ROOT / rel).is_file():
            found[variant] = str(rel).replace("\\", "/")
            continue
        # Backward-compat: old flat layout
        rel_flat = Path("dictionary") / "audio" / variant / f"{slug}{AUDIO_EXT}"
        if (REPO_ROOT / rel_flat).is_file():
            found[variant] = str(rel_flat).replace("\\", "/")
    return found if len(found) > 1 else None


def extract_category_num(text: str) -> str:
    """'29情緒思維（精神性）' -> '29'  (returns empty string if no digit prefix)."""
    if not text:
        return ""
    m = _CATEGORY_PREFIX_RE.match(text)
    return m.group(1) if m else text  # fallback: keep as-is so data isn't lost


def format_id(raw: Any) -> str:
    """Stringify and zero-pad numeric IDs to ID_WIDTH digits."""
    s = normalize_cell(raw)
    if s.isdigit():
        return s.zfill(ID_WIDTH)
    return s


def map_columns(header: tuple) -> dict[str, Any]:
    """Map our logical field names to column indices.
    Returns dict: {logical_name: column_index}.

    Recognized aliases:
      - 1st 'pos' -> 'category';  2nd 'pos' -> 'source' (legacy duplicate-pos layout)
      - 'source', 'souce' (typo), 'variety', 'src', '出處', '出处' -> 'source'
      - 'original_entry', 'ipa' -> 'original'
      - 'remarks', 'note', 'notes', '備考' -> 'remark'
    """
    SOURCE_ALIASES   = {"source", "souce", "variety", "src", "出處", "出处"}
    ORIGINAL_ALIASES = {"original_entry", "original", "ipa", "原表記"}
    REMARK_ALIASES   = {"remark", "remarks", "note", "notes", "備考", "備註"}

    seen_pos: list[int] = []
    col_map: dict[str, int] = {}
    for i, name in enumerate(header):
        if name is None:
            continue
        key = str(name).strip().lower()
        if key == "pos":
            seen_pos.append(i)
            continue
        if key in SOURCE_ALIASES:
            key = "source"
        elif key in ORIGINAL_ALIASES:
            key = "original"
        elif key in REMARK_ALIASES:
            key = "remark"
        col_map[key] = i

    # Fallback for legacy "duplicate pos" layout
    if len(seen_pos) >= 1:
        col_map.setdefault("category", seen_pos[0])
    if len(seen_pos) >= 2:
        col_map.setdefault("source", seen_pos[1])
    return col_map


def read_workbook(xlsx_path: Path) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Return (entries, categories_map)."""
    if not xlsx_path.exists():
        sys.exit(f"ERROR: input file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, keep_vba=xlsx_path.suffix.lower() == ".xlsm")

    # Auto-select the main sheet.
    sheet_name = "basay_dictionary" if "basay_dictionary" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        sys.exit("ERROR: Excel file has no header row.")

    cmap = map_columns(header)
    required = ["id", "basay", "category"]
    missing = [c for c in required if c not in cmap]
    if missing:
        sys.exit("ERROR: required columns missing: " + ", ".join(missing)
                 + f"\n  found columns: {list(cmap.keys())}")

    multi_value_keys = {"zh", "ja", "en"}
    optional_keys = ["zh", "ja", "en", "source", "original", "remark"]

    entries: list[dict[str, Any]] = []
    for row_idx, row in enumerate(rows, start=2):
        if all(v is None or normalize_cell(v) == "" for v in row):
            continue

        record: dict[str, Any] = {}

        # Required: id, basay
        raw_id = row[cmap["id"]] if cmap["id"] < len(row) else None
        record["id"] = format_id(raw_id)
        if not record["id"]:
            print(f"  ! row {row_idx}: missing id, skipped", file=sys.stderr)
            continue

        basay = normalize_cell(row[cmap["basay"]])
        if not basay:
            print(f"  ! row {row_idx} (id={record['id']}): missing basay, skipped",
                  file=sys.stderr)
            continue
        record["basay"] = basay

        # Category: keep only leading digits to save space.
        cat_text = normalize_cell(row[cmap["category"]])
        record["category"] = extract_category_num(cat_text)

        # Optional fields.
        for key in optional_keys:
            if key not in cmap:
                continue
            raw = normalize_cell(row[cmap[key]])
            if key in multi_value_keys:
                vals = split_multivalue(raw)
                if vals:
                    record[key] = vals
            else:
                if raw:
                    record[key] = raw

        # Audio detection.
        slug = derive_slug(record["basay"])
        audio = detect_audio(slug)
        if audio is not None:
            record["audio"] = audio

        entries.append(record)

    # Build the categories.json mapping from the pos_summary sheet if available.
    categories: dict[str, str] = {}
    if "pos_summary" in wb.sheetnames:
        ws2 = wb["pos_summary"]
        for r in ws2.iter_rows(values_only=True):
            if not r or not r[0] or r[0] == "pos":
                continue
            full = normalize_cell(r[0])
            num = extract_category_num(full)
            if num:
                categories[num] = full

    wb.close()
    return entries, categories


def detect_duplicates(entries: list[dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    seen_ids: dict[str, int] = {}
    for i, e in enumerate(entries, start=1):
        if e["id"] in seen_ids:
            warnings.append(f"duplicate id '{e['id']}' at entry #{i}")
        else:
            seen_ids[e["id"]] = i
    # Homograph basay is intentional (handled by ID), not warned here.
    return warnings


def group_by_letter(entries):
    groups: dict[str, list[dict[str, Any]]] = {}
    for e in entries:
        groups.setdefault(initial_letter(e["basay"]), []).append(e)
    for g in groups.values():
        g.sort(key=lambda x: (x["basay"].lower(), x["id"]))
    return groups


def build_search_index(entries):
    """Trimmed view for client-side search. Audio and remark omitted."""
    out = []
    for e in entries:
        item = {
            "id":       e["id"],
            "basay":    e["basay"],
            "category": e.get("category", ""),
        }
        for k in ("zh", "ja", "en"):
            if k in e:
                item[k] = e[k]
        out.append(item)
    return out


def write_json(path: Path, data, indent: int = 2) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=indent)
        f.write("\n")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", "-i", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--no-search-index", action="store_true")
    args = ap.parse_args()

    print(f"Reading {args.input} ...")
    entries, categories = read_workbook(args.input)
    print(f"  loaded {len(entries)} entries; {len(categories)} categories")

    warnings = detect_duplicates(entries)
    for w in warnings:
        print(f"  ! warning: {w}", file=sys.stderr)

    entries_sorted = sorted(entries, key=lambda e: e["id"])
    groups = group_by_letter(entries)
    print(f"  grouped into {len(groups)} letter buckets: "
          + ", ".join(sorted(groups.keys())))

    audio_count = sum(1 for e in entries if "audio" in e)
    if audio_count:
        print(f"  detected audio for {audio_count}/{len(entries)} entries")

    if args.dry_run:
        for letter in sorted(groups):
            print(f"  {letter}.json : {len(groups[letter])} entries")
        return 0

    # Wipe stale per-letter files.
    if ENTRIES_DIR.exists():
        for old in ENTRIES_DIR.glob("*.json"):
            try:
                old.unlink()
            except OSError as e:
                print(f"  ! could not remove {old.name}: {e}", file=sys.stderr)
    ENTRIES_DIR.mkdir(parents=True, exist_ok=True)
    for letter, group in groups.items():
        write_json(ENTRIES_DIR / f"{letter}.json", group)

    id_index = {e["id"]: initial_letter(e["basay"]) for e in entries_sorted}
    write_json(ENTRIES_DIR / "_index.json", id_index)

    if categories:
        write_json(CATEGORIES_JSON, categories)
    write_json(SITE_JSON, entries_sorted)
    if not args.no_search_index:
        write_json(SEARCH_INDEX, build_search_index(entries_sorted))

    print(f"\nWrote: {len(groups)} letter files + _index.json")
    print(f"       {CATEGORIES_JSON}  ({len(categories)} entries)")
    print(f"       {SITE_JSON}")
    if not args.no_search_index:
        print(f"       {SEARCH_INDEX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
